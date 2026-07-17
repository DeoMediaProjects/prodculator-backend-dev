"""Tests for the Excel export endpoint (GET /api/reports/{report_id}/export-excel).

Coverage goals:
- Plan gating: free → 403, professional → 403, producer → 200, studio → 200
- Ownership: another user's report → 403
- Not found: unknown report_id → 404
- Processing: report with no data yet → 422
- Happy path: valid data → valid .xlsx bytes, correct MIME type, all sheets present
- Partial data: missing optional sections → still returns 200, graceful empty sheets
- Filename sanitisation: script title with special chars → safe filename in header
- Excel content: rows written to correct sheets (not just checking bytes length)
"""
import io

import pytest

from app.core.dependencies import get_current_user
from app.modules.auth.schemas import AuthUser
from app.modules.reports.router import get_report_service

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

FULL_REPORT_DATA = {
    "genre": "Drama",
    "tone": "Dark thriller",
    "scale": "Medium",
    "complexity": "High",
    "executiveSummary": {
        "keyInsights": "Strong UK incentive applies.",
        "recommendedTerritory": "United Kingdom",
        "recommendedTerritoryScore": 87,
        "headlineNetBudget": "£2.1M",
        "keyFlags": ["UK EIS relief available", "Budget under HETV threshold"],
    },
    "locationRankings": [
        {
            "name": "London",
            "country": "United Kingdom",
            "score": 87,
            "costEfficiency": 72,
            "crewDepth": 95,
            "infrastructure": 90,
            "incentiveStrength": 85,
            "currencyAdvantage": 68,
            "incentiveReliability": 80,
            "bankabilityLabel": "BANKABLE",
            "rebatePercent": "25%",
            "rebateAmount": "£750,000",
            "reasoning": ["Strong crew base", "High incentive reliability"],
            "keyAdvantages": ["Experienced crew", "Studio infrastructure"],
            "keyRisks": ["High daily rates"],
        }
    ],
    "incentiveEstimates": [
        {
            "territory": "United Kingdom",
            "program": "UK Film Tax Relief",
            "rate": "25%",
            "cap": "No cap",
            "qualifyingSpend": "£3,000,000",
            "estimatedRebate": "£750,000",
            "requirements": ["80% UK spend", "British cultural test"],
            "disclaimer": "Estimate only.",
            "dataSource": "Prodculator DB",
            "lastUpdated": "2026-01-01",
            "bankabilityLabel": "BANKABLE",
        }
    ],
    "financialAnalysis": {
        "budgetScenarios": [
            {
                "territory": "United Kingdom",
                "programme": "UK Film Tax Relief",
                "totalBudget": "£3,000,000",
                "qualifyingSpendPct": "90%",
                "qualifyingSpend": "£2,700,000",
                "atlDeduction": "£270,000",
                "netQualifyingSpend": "£2,430,000",
                "rateGross": "25%",
                "grossRebate": "£607,500",
                "netRebate": "£607,500",
                "netBudget": "£2,392,500",
                "notes": "Assumes full above-the-line deduction.",
            }
        ],
    },
    "weatherLogistics": [
        {
            "territory": "United Kingdom",
            "bestMonths": ["May", "June", "July"],
            "weatherRisk": "Medium",
            "infrastructure": "Excellent transport links.",
            "travelVisa": "No visa for EU crew.",
            "avgTempRange": "10–20°C",
            "shootWindowOverlap": False,
        }
    ],
    "fundingOpportunities": [
        {
            "type": "Fund",
            "name": "BFI Production Fund",
            "genre": ["Drama", "Documentary"],
            "deadline": "2026-06-01",
            "notes": "Up to £500k.",
            "website": "https://www.bfi.org.uk",
        }
    ],
}


class FakeExcelReportService:
    def __init__(self, reports: dict | None = None):
        self._reports: dict = reports or {}

    def get_report(self, report_id: str):
        return self._reports.get(report_id)


def _make_report(
    report_id: str = "rpt-1",
    user_id: str = "user-1",
    script_title: str = "My Script",
    report_data: dict | None = None,
) -> dict:
    return {
        "id": report_id,
        "user_id": user_id,
        "script_title": script_title,
        "status": "completed",
        "report_type": "paid",
        "report_data": report_data,
        "pdf_url": "user-1/rpt-1.pdf",
        "created_at": "2026-01-15T10:00:00Z",
    }


def _make_user(plan: str, user_id: str = "user-1") -> AuthUser:
    return AuthUser(
        id=user_id,
        email=f"{plan}@example.com",
        name="Test User",
        company="Acme",
        role="Producer",
        user_type="paid" if plan != "free" else "free",
        credits_remaining=0,
        plan=plan,
    )


# ── Plan gating ───────────────────────────────────────────────────────────────

@pytest.mark.parametrize("plan,expected_status", [
    ("free", 403),
    ("professional", 403),
    ("producer", 200),
    ("studio", 200),
])
def test_excel_export_plan_gating(client, plan, expected_status):
    """Only Producer and Studio plans can access Excel export."""
    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user(plan)

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == expected_status, (
        f"plan={plan}: expected {expected_status}, got {response.status_code} — {response.text}"
    )


# ── Ownership ─────────────────────────────────────────────────────────────────

def test_excel_export_denies_other_users_report(client):
    """User cannot export a report they don't own."""
    report = _make_report(user_id="owner-user", report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer", user_id="attacker-user")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 403


# ── Not found ─────────────────────────────────────────────────────────────────

def test_excel_export_404_on_unknown_report(client):
    service = FakeExcelReportService({})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/nonexistent/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 404


# ── Still processing ──────────────────────────────────────────────────────────

def test_excel_export_422_when_report_not_yet_complete(client):
    """Reports that are still processing have no report_data — should return 422."""
    report = _make_report(report_data=None)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 422


# ── Happy path ────────────────────────────────────────────────────────────────

def test_excel_export_returns_valid_xlsx_with_correct_mime(client):
    """Producer user gets a valid .xlsx file with the correct MIME type."""
    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200
    assert XLSX_MIME in response.headers["content-type"]
    assert len(response.content) > 1000, "Response body too small to be a real workbook"


def test_excel_export_content_disposition_header(client):
    """Content-Disposition header should contain the script title."""
    report = _make_report(script_title="My Great Film", report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200
    cd = response.headers.get("content-disposition", "")
    assert "My Great Film" in cd


def test_excel_export_filename_sanitises_special_chars(client):
    """Special characters in script title should be stripped from the filename."""
    report = _make_report(script_title="Film: The 100% Truth (Draft)", report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200
    cd = response.headers.get("content-disposition", "")
    # Colons and percent signs should be gone
    assert ":" not in cd
    assert "%" not in cd


# ── Sheet content ─────────────────────────────────────────────────────────────

def test_excel_export_all_expected_sheets_present(client):
    """Workbook must contain the 8 expected sheets."""
    from openpyxl import load_workbook

    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.content))
    sheet_names = wb.sheetnames
    expected_sheets = [
        "Summary",
        "Territory Rankings",
        "Tax Incentives",
        "Financial Analysis",
        "Comparable Productions",
        "Weather & Logistics",
        "Funding & Festivals",
    ]
    for sheet in expected_sheets:
        assert sheet in sheet_names, f"Missing sheet: {sheet}"


def test_excel_export_territory_rankings_sheet_has_correct_data(client):
    """Territory Rankings sheet must contain territory name in data rows."""
    from openpyxl import load_workbook

    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Territory Rankings"]

    # Row 1 is header — row 2 is first data row
    cell_values = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert 1 in cell_values, "Rank column missing"
    assert "London" in cell_values, "Territory name missing"
    assert "United Kingdom" in cell_values, "Country missing"
    assert 87 in cell_values, "Score missing"


def test_excel_export_summary_sheet_contains_script_title(client):
    """Summary sheet must list the script title."""
    from openpyxl import load_workbook

    report = _make_report(script_title="Test Feature Film", report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Summary"]

    all_values = [ws.cell(row=r, column=2).value for r in range(1, 15)]
    assert "Test Feature Film" in all_values


def test_excel_export_incentives_sheet_has_territory_and_rebate(client):
    """Tax Incentives sheet must have the territory name and estimated rebate."""
    from openpyxl import load_workbook

    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Tax Incentives"]

    row2_values = [ws.cell(row=2, column=c).value for c in range(1, 10)]
    assert "United Kingdom" in row2_values
    assert "£750,000" in row2_values


# ── Partial data / edge cases ─────────────────────────────────────────────────

def test_excel_export_handles_missing_optional_sections_gracefully(client):
    """Report with only minimal data (no optional sections) should still return 200."""
    minimal_data = {
        "genre": "Documentary",
        "tone": "Observational",
        "scale": "Small",
        "complexity": "Low",
        "locationRankings": [],
        "incentiveEstimates": [],
        "comparables": [],
        "weatherLogistics": [],
        "fundingOpportunities": [],
    }
    report = _make_report(report_data=minimal_data)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("producer")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200


def test_excel_export_studio_plan_also_allowed(client):
    """Studio users should also receive 200 — they inherit producer permissions."""
    report = _make_report(report_data=FULL_REPORT_DATA)
    service = FakeExcelReportService({"rpt-1": report})
    user = _make_user("studio")

    client.app.dependency_overrides[get_current_user] = lambda: user
    client.app.dependency_overrides[get_report_service] = lambda: service

    response = client.get("/api/reports/rpt-1/export-excel", headers={"Authorization": "Bearer token"})
    assert response.status_code == 200


# ── Direct unit test of excel_service ────────────────────────────────────────

def test_build_excel_workbook_produces_valid_xlsx_bytes():
    """Unit test the service directly — no HTTP layer needed."""
    from openpyxl import load_workbook
    from app.modules.reports.excel_service import build_excel_workbook

    report = _make_report(report_data=FULL_REPORT_DATA)
    raw_bytes = build_excel_workbook(report)

    assert isinstance(raw_bytes, bytes)
    assert len(raw_bytes) > 0

    wb = load_workbook(io.BytesIO(raw_bytes))
    assert "Summary" in wb.sheetnames
    assert "Territory Rankings" in wb.sheetnames


def test_build_excel_workbook_with_empty_report_data():
    """Empty report_data should produce a workbook without crashing."""
    from openpyxl import load_workbook
    from app.modules.reports.excel_service import build_excel_workbook

    report = _make_report(report_data={})
    raw_bytes = build_excel_workbook(report)

    wb = load_workbook(io.BytesIO(raw_bytes))
    # Summary sheet always exists
    assert "Summary" in wb.sheetnames
    # Territory Rankings created but empty (just headers)
    assert "Territory Rankings" in wb.sheetnames
