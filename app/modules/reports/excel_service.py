"""Excel export service for report data.

Produces a multi-sheet .xlsx workbook from a completed report's JSON data.
Each major report section gets its own worksheet for easy filtering/analysis.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)

# Gold accent used across sheets to match brand
_GOLD = "D4AF37"
_BLACK = "000000"
_WHITE = "FFFFFF"
_LIGHT_GREY = "F5F5F5"
_GREEN = "4CAF50"


def build_excel_workbook(report: dict) -> bytes:
    """Build a .xlsx workbook from a report dict and return raw bytes.

    Raises ImportError if openpyxl is not installed.
    Returns empty workbook bytes if report_data is absent.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    data: dict = report.get("report_data") or {}
    script_title: str = report.get("script_title", "Untitled")
    created_at: str = str(report.get("created_at", ""))[:10]

    def _header_font() -> Font:
        return Font(bold=True, color=_BLACK, size=11)

    def _gold_fill() -> PatternFill:
        return PatternFill("solid", fgColor=_GOLD)

    def _grey_fill() -> PatternFill:
        return PatternFill("solid", fgColor=_LIGHT_GREY)

    def _thin_border() -> Border:
        side = Side(style="thin", color="CCCCCC")
        return Border(left=side, right=side, top=side, bottom=side)

    def _write_header_row(ws: Any, headers: list[str], row: int = 1) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _header_font()
            cell.fill = _gold_fill()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border = _thin_border()

    def _write_data_row(ws: Any, values: list[Any], row: int, shade: bool = False) -> None:
        fill = _grey_fill() if shade else None
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

    def _autofit_columns(ws: Any, min_width: int = 12, max_width: int = 50) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.freeze_panes = "B1"

    exec_summary: dict = data.get("executiveSummary") or {}
    summary_rows = [
        ("Script Title", script_title),
        ("Report Date", created_at),
        ("Genre", data.get("genre", "")),
        ("Tone", data.get("tone", "")),
        ("Scale", data.get("scale", "")),
        ("Complexity", data.get("complexity", "")),
        ("Recommended Territory", exec_summary.get("recommendedTerritory", "")),
        ("Recommended Score", exec_summary.get("recommendedTerritoryScore", "")),
        ("Headline Net Budget", exec_summary.get("headlineNetBudget", "")),
        ("Key Insights", exec_summary.get("keyInsights", "")),
    ]
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 55
    for row_idx, (label, value) in enumerate(summary_rows, 1):
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.fill = _gold_fill()
        label_cell.border = _thin_border()
        label_cell.alignment = Alignment(vertical="center")
        value_cell = ws_summary.cell(row=row_idx, column=2, value=str(value) if value is not None else "")
        value_cell.border = _thin_border()
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Key flags
    key_flags: list[str] = exec_summary.get("keyFlags") or []
    if key_flags:
        start = len(summary_rows) + 2
        ws_summary.cell(row=start, column=1, value="Key Flags").font = Font(bold=True, size=11)
        for fi, flag in enumerate(key_flags):
            ws_summary.cell(row=start + 1 + fi, column=1, value=f"• {flag}")

    # ── Sheet 2: Territory Rankings ──────────────────────────────────────────
    ws_terr = wb.create_sheet("Territory Rankings")
    ws_terr.freeze_panes = "A2"

    terr_headers = [
        "Rank", "Territory", "Country", "Overall Score",
        "Cost Efficiency", "Crew Depth", "Infrastructure",
        "Incentive Strength", "Currency Advantage", "Incentive Reliability",
        "Bankability", "Rebate %", "Est. Rebate Amount",
        "Cultural Test Likelihood", "Admin Complexity", "Payment Speed",
        "Key Advantages", "Key Risks",
    ]
    _write_header_row(ws_terr, terr_headers)

    location_rankings: list[dict] = data.get("locationRankings") or []
    for rank_idx, loc in enumerate(location_rankings, 1):
        _write_data_row(
            ws_terr,
            [
                rank_idx,
                loc.get("name", ""),
                loc.get("country", ""),
                loc.get("score", ""),
                loc.get("costEfficiency", ""),
                loc.get("crewDepth", ""),
                loc.get("infrastructure", ""),
                loc.get("incentiveStrength", ""),
                loc.get("currencyAdvantage", ""),
                loc.get("incentiveReliability", ""),
                loc.get("bankabilityLabel", ""),
                loc.get("rebatePercent", ""),
                loc.get("rebateAmount", ""),
                loc.get("culturalTestLikelihood", ""),
                loc.get("adminComplexity", ""),
                loc.get("paymentSpeed", ""),
                "\n".join(loc.get("keyAdvantages") or []),
                "\n".join(loc.get("keyRisks") or []),
            ],
            row=rank_idx + 1,
            shade=rank_idx % 2 == 0,
        )

    _autofit_columns(ws_terr)

    # ── Sheet 3: Tax Incentives ──────────────────────────────────────────────
    ws_inc = wb.create_sheet("Tax Incentives")
    ws_inc.freeze_panes = "A2"

    inc_headers = [
        "Territory", "Programme", "Rate", "Cap",
        "Qualifying Spend", "Estimated Rebate",
        "Rate Type", "Bankability", "Payment Speed",
        "Scope", "Stackable With", "Eligibility Status", "Eligibility Note",
        "Expiry Date", "Data Freshness", "Data Source", "Last Updated",
        "Requirements", "Disclaimer", "Warnings",
    ]
    _write_header_row(ws_inc, inc_headers)

    incentive_estimates: list[dict] = data.get("incentiveEstimates") or []
    for row_idx, inc in enumerate(incentive_estimates, 2):
        _write_data_row(
            ws_inc,
            [
                inc.get("territory", ""),
                inc.get("program", ""),
                inc.get("rate", ""),
                inc.get("cap", ""),
                inc.get("qualifyingSpend", ""),
                inc.get("estimatedRebate", ""),
                inc.get("rateType", ""),
                inc.get("bankabilityLabel", ""),
                inc.get("paymentSpeed", ""),
                inc.get("scope", ""),
                ", ".join(inc.get("stackableWith") or []),
                inc.get("eligibilityStatus", ""),
                inc.get("eligibilityNote", ""),
                inc.get("expiryDate", ""),
                inc.get("dataFreshness", ""),
                inc.get("dataSource", ""),
                inc.get("lastUpdated", ""),
                "\n".join(inc.get("requirements") or []),
                inc.get("disclaimer", ""),
                "\n".join(inc.get("warnings") or []),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_inc)

    # ── Sheet 4: Financial Analysis ──────────────────────────────────────────
    ws_fin = wb.create_sheet("Financial Analysis")
    ws_fin.freeze_panes = "A2"

    fin_headers = [
        "Territory", "Programme", "Total Budget", "Qualifying Spend %",
        "Qualifying Spend", "ATL Deduction", "ATL Deduction %",
        "Net Qualifying Spend", "Gross Rate", "Net Rate",
        "Gross Rebate", "Net Rebate", "Net Budget After Rebate",
        "Local Spend (legacy)", "Rebate Rate (legacy)", "Notes",
    ]
    _write_header_row(ws_fin, fin_headers)

    financial_analysis: dict = data.get("financialAnalysis") or {}
    budget_scenarios: list[dict] = financial_analysis.get("budgetScenarios") or []
    for row_idx, scenario in enumerate(budget_scenarios, 2):
        _write_data_row(
            ws_fin,
            [
                scenario.get("territory", ""),
                scenario.get("programme", ""),
                scenario.get("totalBudget", ""),
                scenario.get("qualifyingSpendPct", ""),
                scenario.get("qualifyingSpend", ""),
                scenario.get("atlDeduction", ""),
                scenario.get("atlDeductionPct", ""),
                scenario.get("netQualifyingSpend", ""),
                scenario.get("rateGross", ""),
                scenario.get("rateNet", ""),
                scenario.get("grossRebate", ""),
                scenario.get("netRebate", ""),
                scenario.get("netBudget", ""),
                scenario.get("localSpend", ""),
                scenario.get("rebateRate", ""),
                scenario.get("notes", ""),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_fin)

    # ── Sheet 4b: Crew Cost Comparison (sub-section of financialAnalysis) ────
    crew_cost_rows: list[dict] = financial_analysis.get("crewCostComparison") or []
    if crew_cost_rows:
        ws_crew_cost = wb.create_sheet("Crew Cost Comparison")
        ws_crew_cost.freeze_panes = "A2"

        # Territory names are dynamic keys — collect all unique ones
        all_territories: list[str] = []
        seen: set[str] = set()
        for row in crew_cost_rows:
            for t in (row.get("territories") or {}).keys():
                if t not in seen:
                    all_territories.append(t)
                    seen.add(t)

        crew_cost_headers = ["Role"] + all_territories
        _write_header_row(ws_crew_cost, crew_cost_headers)

        for row_idx, row in enumerate(crew_cost_rows, 2):
            territories_map: dict = row.get("territories") or {}
            _write_data_row(
                ws_crew_cost,
                [row.get("role", "")] + [territories_map.get(t, "") for t in all_territories],
                row=row_idx,
                shade=row_idx % 2 == 0,
            )

        _autofit_columns(ws_crew_cost)

    # ── Sheet 5: Crew Insights ───────────────────────────────────────────────
    ws_crew = wb.create_sheet("Crew Insights")
    ws_crew.freeze_panes = "A2"

    crew_headers = [
        "Territory", "Availability", "Cost vs USD", "Quality Rating (1-5)",
        "Specialties", "Tradeoff Note",
        "Currency", "FX Rate", "FX Date", "Data Source",
    ]
    _write_header_row(ws_crew, crew_headers)

    crew_insights: list[dict] = data.get("crewInsights") or []
    for row_idx, crew in enumerate(crew_insights, 2):
        _write_data_row(
            ws_crew,
            [
                crew.get("territory", ""),
                crew.get("availability", ""),
                crew.get("costVsUSD", ""),
                crew.get("qualityRating", ""),
                ", ".join(crew.get("specialties") or []),
                crew.get("tradeoff", ""),
                crew.get("currency", ""),
                crew.get("fxRate", ""),
                crew.get("fxDate", ""),
                crew.get("dataSource", ""),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_crew)

    # ── Sheet 6: Comparable Productions ─────────────────────────────────────
    ws_comp = wb.create_sheet("Comparable Productions")
    ws_comp.freeze_panes = "A2"

    comp_headers = [
        "Title", "Genre", "Budget Range", "Visual Scale",
        "Location", "Year", "Source", "Relevance",
    ]
    _write_header_row(ws_comp, comp_headers)

    comparables: list[dict] = data.get("comparables") or []
    for row_idx, comp in enumerate(comparables, 2):
        _write_data_row(
            ws_comp,
            [
                comp.get("title", ""),
                comp.get("genre", ""),
                comp.get("budgetRange", ""),
                comp.get("visualScale", ""),
                comp.get("location", ""),
                comp.get("year", ""),
                comp.get("source", ""),
                comp.get("relevanceDescription", ""),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_comp)

    # ── Sheet 7: Weather & Logistics ─────────────────────────────────────────
    ws_weather = wb.create_sheet("Weather & Logistics")
    ws_weather.freeze_panes = "A2"

    weather_headers = [
        "Territory", "Weather Risk", "Best Months",
        "Avg Temp Range", "Avg Rainfall", "Daylight Hours",
        "Infrastructure", "Travel & Visa",
        "Seasonal Considerations", "Shoot Window Overlap",
        "Shoot Window Risk", "Estimated Delay Days", "Contingency Budget",
    ]
    _write_header_row(ws_weather, weather_headers)

    weather_logistics: list[dict] = data.get("weatherLogistics") or []
    for row_idx, weather in enumerate(weather_logistics, 2):
        _write_data_row(
            ws_weather,
            [
                weather.get("territory", ""),
                weather.get("weatherRisk", ""),
                ", ".join(weather.get("bestMonths") or []),
                weather.get("avgTempRange", ""),
                weather.get("avgRainfall", ""),
                weather.get("daylightHours", ""),
                weather.get("infrastructure", ""),
                weather.get("travelVisa", ""),
                weather.get("seasonalConsiderations", ""),
                "Yes" if weather.get("shootWindowOverlap") else ("No" if weather.get("shootWindowOverlap") is False else ""),
                weather.get("shootWindowRisk", ""),
                weather.get("estimatedDelayDays", ""),
                weather.get("contingencyBudget", ""),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_weather)

    # ── Sheet 8: Funding & Festivals ─────────────────────────────────────────
    ws_fund = wb.create_sheet("Funding & Festivals")
    ws_fund.freeze_panes = "A2"

    fund_headers = [
        "Name", "Type", "Genre", "Deadline", "Tier", "Notes", "Website",
    ]
    _write_header_row(ws_fund, fund_headers)

    funding_opportunities: list[dict] = data.get("fundingOpportunities") or []
    for row_idx, opp in enumerate(funding_opportunities, 2):
        _write_data_row(
            ws_fund,
            [
                opp.get("name", ""),
                opp.get("type", ""),
                ", ".join(opp.get("genre") or []),
                opp.get("deadline", ""),
                opp.get("tier", ""),
                opp.get("notes", ""),
                opp.get("website", ""),
            ],
            row=row_idx,
            shade=row_idx % 2 == 0,
        )

    _autofit_columns(ws_fund)

    # ── Write to bytes ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
